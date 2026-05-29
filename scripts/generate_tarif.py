#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate_tarif.py — Génère tarif.json à partir du fichier maître TarifVinom_Master.xlsx

Le fichier maître a une feuille "Tarif" entièrement structurée (1 ligne = 1 référence),
avec des colonnes nommées explicites. Le script lit ces colonnes directement —
plus aucune heuristique ni matching nécessaire, tout est dans le fichier.

Usage :
    python scripts/generate_tarif.py --excel data/TarifVinom_Master.xlsx --output tarif.json

Colonnes attendues dans la feuille "Tarif" :
    Code_Vinistoria, Statut, Region, Sous_region, Couleur, Appellation,
    Cuvee, Domaine, Label, Millesime, Contenance_cl, Tarif_HT, Stock,
    Arrivage_prevu, Exclusivite, Date_ajout, Notes_internes, Pepite

Conçu pour tourner en CI (GitHub Actions) sans supervision :
- codes de sortie explicites (0 = OK, != 0 = échec)
- garde-fou : refuse de publier un catalogue manifestement cassé.
"""

import argparse
import json
import sys
from datetime import date

import openpyxl


# ===================================================================
# Constantes métier
# ===================================================================

REGION_ORDER = [
    "Île-de-France", "Jura", "Alsace", "Bordeaux", "Beaujolais",
    "Bourgogne", "Rhône", "Loire", "Provence-Corse", "Occitanie",
    "Reste du Monde", "Bulles", "Spiritueux",
]

# Colonnes attendues (nom -> on les retrouve par leur intitulé, pas par position,
# pour rester robuste si l'ordre change un jour)
COLONNES = [
    "Code_Vinistoria", "Statut", "Region", "Sous_region", "Couleur",
    "Appellation", "Cuvee", "Domaine", "Label", "Millesime",
    "Contenance_cl", "Tarif_HT", "Stock", "Arrivage_prevu",
    "Exclusivite", "Date_ajout", "Notes_internes", "Pepite",
]

# Normalisation des couleurs vers les clés utilisées par la page web
COULEUR_NORM = {
    "rouge": "rouge", "rge": "rouge",
    "blanc": "blanc", "blc": "blanc",
    "rosé": "rose", "rose": "rose", "ros": "rose",
    "bulles": "bulles", "pet": "bulles",
    "spiritueux": "spiritueux", "spi": "spiritueux",
}

# Garde-fou : on refuse de publier si trop peu de références (Excel cassé / mauvaise feuille)
# Le catalogue compte ~500 références ; en dessous de 400, on suspecte un problème.
MIN_REFERENCES = 400

EDITO_DEFAUT = (
    "À l'approche des beaux jours, votre carte des vins retrouve naturellement "
    "des allures ensoleillées. C'est la saison idéale pour redécouvrir nos rosés "
    "de Provence, nos Chablis frais et minéraux, et l'ensemble de notre sélection "
    "actuelle. Bonne lecture et bonnes dégustations."
)


# ===================================================================
# Utilitaires
# ===================================================================

def log(msg):
    print(f"[generate_tarif] {msg}", flush=True)


def fail(msg, code=1):
    print(f"[generate_tarif] ERREUR : {msg}", file=sys.stderr, flush=True)
    sys.exit(code)


def cell_str(v):
    """Renvoie une chaîne propre depuis une cellule (gère None, nombres, espaces)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def deduce_couleur(couleur_brute, region, cuvee, appellation):
    """La colonne Couleur fait foi. Si vide, on déduit a minima par région
    (Bulles/Spiritueux) puis par quelques mots-clés simples."""
    c = (couleur_brute or "").strip().lower()
    if c in COULEUR_NORM:
        return COULEUR_NORM[c]

    if region == "Bulles":
        return "bulles"
    if region == "Spiritueux":
        return "spiritueux"

    txt = f" {cuvee.lower()} {appellation.lower()} "
    if any(k in txt for k in ["rosé", " rose ", " gris "]):
        return "rose"
    if any(k in txt for k in ["chardonnay", "sauvignon", "chenin", "viognier",
                              "riesling", "blanc", "aligoté", "muscat",
                              "gewurtz", "savagnin", "pinot gris", "moelleux"]):
        return "blanc"
    if any(k in txt for k in ["pinot noir", "merlot", "cabernet", "gamay",
                              "syrah", "malbec", "rouge", " noir", "grenache"]):
        return "rouge"
    # défaut neutre : rouge (dominante du catalogue)
    return "rouge"


def label_tags(label):
    """Convertit le libellé de label en étiquettes filtrables.
    DEMETER = BIODYNAMIE (consigne métier)."""
    if not label:
        return []
    tags, l = [], label.strip().lower()
    if "bio" in l:
        tags.append("biodynamie" if ("demeter" in l or "biodynam" in l) else "bio")
    if ("demeter" in l or "biodynamie" in l) and "biodynamie" not in tags:
        tags.append("biodynamie")
    if "no so2" in l or "nature" in l or "sans sulfites" in l:
        tags.append("sans_so2")
    if "hve" in l:
        tags.append("hve")
    if "vegan" in l:
        tags.append("vegan")
    if "régénérative" in l:
        tags.append("regen")
    return tags


def format_from_contenance(contenance):
    """Déduit le tag format + libellé depuis la contenance en cl."""
    c = (contenance or "").strip().replace(",", ".")
    try:
        val = float(c)
    except ValueError:
        return "standard", "75cl"
    if val >= 1000:
        return "bib", f"BIB {int(val/100)}L"
    if val == 500:
        return "bib", "BIB 5L"
    if val == 300:
        return "jeroboam", "300cl"
    if val == 150:
        return "magnum", "150cl"
    if val == 37.5:
        return "demi", "37,5cl"
    if val == 70:
        return "standard", "70cl"
    return "standard", "75cl"


# ===================================================================
# Lecture du fichier maître
# ===================================================================

def read_master(path):
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        fail(f"impossible d'ouvrir le fichier Excel : {e}")

    if "Tarif" not in wb.sheetnames:
        fail(f"feuille 'Tarif' absente. Feuilles trouvées : {wb.sheetnames}")

    ws = wb["Tarif"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        fail("la feuille 'Tarif' est vide.")

    headers = [cell_str(h) for h in rows[0]]
    # Index des colonnes par nom (robuste au changement d'ordre)
    idx = {}
    for col in COLONNES:
        if col in headers:
            idx[col] = headers.index(col)
    manquantes = [c for c in ("Region", "Cuvee", "Tarif_HT") if c not in idx]
    if manquantes:
        fail(f"colonnes essentielles manquantes : {manquantes}. "
             f"En-têtes trouvés : {headers}")

    def get(row, col):
        i = idx.get(col)
        return cell_str(row[i]) if i is not None and i < len(row) else ""

    # Lire l'édito depuis la feuille Edito si présente
    edito = EDITO_DEFAUT
    edition_label = ""
    if "Edito" in wb.sheetnames:
        ed = list(wb["Edito"].iter_rows(values_only=True))
        if len(ed) >= 2 and ed[1]:
            # colonnes attendues : Mois | Titre | Texte
            if len(ed[1]) >= 3 and ed[1][2]:
                edito = cell_str(ed[1][2])

    items = []
    for row in rows[1:]:
        cuvee = get(row, "Cuvee")
        domaine = get(row, "Domaine")
        prix_str = get(row, "Tarif_HT").replace(",", ".")
        # Une référence est valide si elle a un nom (cuvée OU domaine) et un prix.
        # Beaucoup de vins n'ont pas de cuvée distincte : leur nom = le domaine
        # (ex. "Château Cheval Noir" en Saint-Émilion).
        if (not cuvee and not domaine) or not prix_str:
            continue
        try:
            prix = float(prix_str)
        except ValueError:
            continue
        if prix <= 0:
            continue

        region = get(row, "Region")
        appellation = get(row, "Appellation")
        couleur = deduce_couleur(get(row, "Couleur"), region, cuvee or domaine, appellation)
        fmt, fmt_label = format_from_contenance(get(row, "Contenance_cl"))
        code = get(row, "Code_Vinistoria")
        stock_str = get(row, "Stock")
        try:
            stock = float(stock_str) if stock_str else None
        except ValueError:
            stock = None

        # Si pas de cuvée, on affiche le domaine comme libellé principal
        # et on n'a pas de "domaine" séparé à répéter.
        if cuvee:
            cuvee_aff, domaine_aff = cuvee, domaine
        else:
            cuvee_aff, domaine_aff = domaine, ""

        items.append({
            "id": code or f"X{len(items):04d}",
            "cuvee": cuvee_aff,
            "domaine": domaine_aff,
            "appellation": appellation,
            "millesime": get(row, "Millesime"),
            "label_display": get(row, "Label"),
            "label_tags": label_tags(get(row, "Label")),
            "prix": prix,
            "region": region,
            "sous_region": get(row, "Sous_region"),
            "couleur": couleur,
            "format": fmt,
            "format_label": fmt_label,
            "nouveau": get(row, "Statut").lower() == "nouveau",
            "pepite": get(row, "Pepite").lower() in ("oui", "x", "true", "1"),
            "exclusivite": get(row, "Exclusivite").lower() in ("oui", "x", "true", "1"),
            "stock": stock,
        })

    log(f"{len(items)} références lues dans la feuille 'Tarif'.")
    return items, edition_label, edito


# ===================================================================
# Main
# ===================================================================

def main():
    ap = argparse.ArgumentParser(description="Génère tarif.json depuis TarifVinom_Master.xlsx")
    ap.add_argument("--excel", required=True, help="Fichier maître (.xlsx)")
    ap.add_argument("--output", default="tarif.json", help="Fichier de sortie JSON")
    ap.add_argument("--edition", default="", help="Libellé d'édition (ex. 'Mai 2026')")
    args = ap.parse_args()

    log(f"Lecture de {args.excel}")
    items, edition_from_file, edito = read_master(args.excel)

    if len(items) < MIN_REFERENCES:
        fail(f"seulement {len(items)} références (seuil minimal : {MIN_REFERENCES}). "
             f"Publication annulée pour éviter un catalogue incomplet. "
             f"Vérifiez le fichier (bonne feuille 'Tarif' ? données présentes ?).")

    # Libellé d'édition : argument > feuille Edito > mois courant
    if args.edition:
        edition_label = args.edition
    elif edition_from_file:
        edition_label = edition_from_file
    else:
        mois = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet",
                "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        t = date.today()
        edition_label = f"{mois[t.month - 1]} {t.year}"

    # Stats de contrôle (affichées dans les logs du workflow)
    from collections import Counter
    couleurs = Counter(i["couleur"] for i in items)
    log(f"Répartition couleurs : {dict(couleurs)}")
    log(f"Nouveautés : {sum(1 for i in items if i['nouveau'])} | "
        f"Pépites : {sum(1 for i in items if i['pepite'])} | "
        f"Avec stock : {sum(1 for i in items if i['stock'] is not None)}")

    data = {
        "meta": {
            "total": len(items),
            "date_maj": date.today().isoformat(),
            "edition": edition_label,
            "edito": edito,
        },
        "items": items,
    }

    try:
        with open(args.output, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    except Exception as e:
        fail(f"écriture de {args.output} impossible : {e}")

    log(f"OK — {len(items)} références écrites dans {args.output} (édition {edition_label}).")
    sys.exit(0)


if __name__ == "__main__":
    main()
